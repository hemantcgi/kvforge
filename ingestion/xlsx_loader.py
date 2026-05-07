"""XLSX spreadsheet loader that chunks rows from multiple sheets with header metadata.

Depends on ``openpyxl`` for reading Excel files. Install with ``pip install openpyxl``.
"""
import hashlib
from pathlib import Path
import openpyxl


def _section_hash(text: str) -> str:
    """Compute SHA-256 hash of text."""
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


class XlsxLoader:
    """Load an XLSX file and split rows into chunks with sheet and header metadata.

    Each sheet's first row is treated as column headers. Remaining rows are grouped
    into chunks of ``rows_per_chunk`` data rows each. Each chunk is formatted as
    rows with "Header: value" format (skipping empty values), joined by newlines.

    Args:
        rows_per_chunk: Number of data rows per chunk. Default is 50.
    """

    def __init__(self, rows_per_chunk: int = 50):
        self.rows_per_chunk = rows_per_chunk

    def load(self, source: str) -> list[dict]:
        """Read an XLSX file and return its rows as a list of chunk dicts.

        Args:
            source: Path to the XLSX file.

        Returns:
            List of dicts with the shape::

                {
                    "text": str,
                    "metadata": {
                        "sheet_name": str,           # Name of the sheet
                        "column_headers": list[str], # Header row values
                        "row_range": {               # 1-based row numbers
                            "start": int,            # First data row (row 2 = index 1)
                            "end": int,              # Last data row (inclusive)
                        },
                        "source": str,               # filename (not full path)
                        "section_hash": str,         # SHA-256 of chunk text
                        "chunk_id": int,             # global 0-indexed chunk counter
                    }
                }
        """
        path = Path(source)
        wb = openpyxl.load_workbook(str(path))
        chunks = []
        chunk_id = 0

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))

                # Skip empty sheets
                if not rows:
                    continue

                # First row is headers, rest are data
                headers = rows[0]
                data_rows = rows[1:]

                # Skip sheets with only headers and no data
                if not data_rows:
                    continue

                # Process data rows in chunks
                for chunk_start in range(0, len(data_rows), self.rows_per_chunk):
                    chunk_end = min(chunk_start + self.rows_per_chunk, len(data_rows))
                    chunk_data_rows = data_rows[chunk_start:chunk_end]

                    # Format each row as "Header: value | Header: value"
                    formatted_rows = []
                    for row in chunk_data_rows:
                        formatted_cells = []
                        for header, value in zip(headers, row):
                            # Skip None/empty values
                            if value is not None and str(value).strip():
                                formatted_cells.append(f"{header}: {value}")
                        if formatted_cells:  # Only add non-empty rows
                            formatted_rows.append(" | ".join(formatted_cells))

                    # Skip if no valid rows in this chunk
                    if not formatted_rows:
                        continue

                    chunk_text = "\n".join(formatted_rows)
                    section_hash = _section_hash(chunk_text)

                    # Calculate 1-based row numbers (row 1 is header)
                    row_start = chunk_start + 2  # +1 for 0->1 indexing, +1 for header
                    row_end = chunk_end + 1      # +1 for 0->1 indexing, +1 for header

                    chunks.append({
                        "text": chunk_text,
                        "metadata": {
                            "sheet_name": sheet_name,
                            "column_headers": list(headers),
                            "row_range": {
                                "start": row_start,
                                "end": row_end,
                            },
                            "source": path.name,
                            "section_hash": section_hash,
                            "chunk_id": chunk_id,
                        },
                    })
                    chunk_id += 1

        finally:
            wb.close()

        return chunks
